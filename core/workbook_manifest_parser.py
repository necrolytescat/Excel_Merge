"""只读提取工作簿 main Sheet 中的 CSV 导出清单。"""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import posixpath
import re
from typing import Any, Iterable
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from core.m2_errors import M2ProcessingError


_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_CELL_REF_RE = re.compile(r"^([A-Z]+)")


@dataclass(frozen=True)
class ManifestEntry:
    sheet_name: str
    tbx_name: str
    is_export: str
    row_number: int


@dataclass(frozen=True)
class WorkbookManifest:
    entries: tuple[ManifestEntry, ...]
    parser: str


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    return str(value)


def _find_sheet_name(sheet_names: Iterable[str], configured: str) -> str:
    names = list(sheet_names)
    if configured in names:
        return configured
    matches = [name for name in names if name.casefold() == configured.casefold()]
    if len(matches) == 1:
        return matches[0]
    raise M2ProcessingError(
        "M2_MANIFEST_SHEET_MISSING",
        "manifest_parse",
        f"工作簿缺少 {configured} Sheet",
        sheet_name=configured,
    )


def _select_manifest_table_ref(
    table_records: list[tuple[str, set[str]]],
    *,
    sheet_name: str,
    required_fields: tuple[str, str, str],
) -> str | None:
    if not table_records:
        return None
    required = set(required_fields)
    candidates = [ref for ref, fields in table_records if required <= fields]
    if len(candidates) != 1:
        raise M2ProcessingError(
            "M2_MANIFEST_FIELD_MISSING",
            "manifest_parse",
            "main Sheet 必须包含唯一的 manifest Excel Table",
            sheet_name=sheet_name,
            details={
                "candidate_table_count": len(candidates),
                "table_refs": candidates,
            },
        )
    return candidates[0]


def _parse_manifest_rows(
    rows: list[tuple[int, list[Any]]],
    *,
    sheet_name: str,
    sheet_field: str,
    csv_name_field: str,
    export_flag_field: str,
    parser: str,
) -> WorkbookManifest:
    header_matches: list[tuple[int, dict[str, int]]] = []
    required_fields = (sheet_field, csv_name_field, export_flag_field)
    for row_number, values in rows:
        text_values = [_cell_text(value) for value in values]
        positions: dict[str, int] = {}
        for index, value in enumerate(text_values):
            if value and value not in positions:
                positions[value] = index
        if all(field in positions for field in required_fields):
            header_matches.append((row_number, positions))

    if len(header_matches) != 1:
        raise M2ProcessingError(
            "M2_MANIFEST_FIELD_MISSING",
            "manifest_parse",
            "main Sheet 必须包含唯一的 sheetName/tbxName/isExport 表头行",
            sheet_name=sheet_name,
            details={"header_match_count": len(header_matches)},
        )

    header_row, positions = header_matches[0]
    entries: list[ManifestEntry] = []
    seen_sheets: dict[str, int] = {}
    for row_number, values in rows:
        if row_number <= header_row:
            continue
        text_values = [_cell_text(value) for value in values]

        def get(field: str) -> str:
            index = positions.get(field)
            return text_values[index] if index is not None and index < len(text_values) else ""

        if get(export_flag_field) != "1":
            continue
        logical_name = get(sheet_field)
        tbx_name = get(csv_name_field)
        missing_fields = [
            field
            for field, value in (
                (sheet_field, logical_name),
                (csv_name_field, tbx_name),
            )
            if value == ""
        ]
        if missing_fields:
            raise M2ProcessingError(
                "M2_MANIFEST_FIELD_MISSING",
                "manifest_parse",
                "main 导出清单行缺少 sheetName 或 tbxName",
                sheet_name=sheet_name,
                details={"row": row_number, "fields": missing_fields},
            )
        if logical_name in seen_sheets:
            raise M2ProcessingError(
                "M2_MANIFEST_DUPLICATE_SHEET",
                "manifest_parse",
                f"main 清单存在重复 sheetName：{logical_name}",
                sheet_name=logical_name,
                details={"rows": [seen_sheets[logical_name], row_number]},
            )
        seen_sheets[logical_name] = row_number
        entries.append(
            ManifestEntry(
                sheet_name=logical_name,
                tbx_name=tbx_name,
                is_export="1",
                row_number=row_number,
            )
        )
    return WorkbookManifest(entries=tuple(entries), parser=parser)


def _parse_with_openpyxl(
    raw: bytes,
    *,
    sheet_name: str,
    sheet_field: str,
    csv_name_field: str,
    export_flag_field: str,
) -> WorkbookManifest:
    workbook = load_workbook(
        BytesIO(raw),
        read_only=False,
        data_only=True,
        keep_vba=False,
    )
    try:
        actual_name = _find_sheet_name(workbook.sheetnames, sheet_name)
        sheet = workbook[actual_name]
        table_records: list[tuple[str, set[str]]] = []
        for table in sheet.tables.values():
            min_column, min_row, max_column, _ = range_boundaries(table.ref)
            header_values = next(
                sheet.iter_rows(
                    min_row=min_row,
                    max_row=min_row,
                    min_col=min_column,
                    max_col=max_column,
                    values_only=True,
                )
            )
            table_records.append(
                (table.ref, {_cell_text(value) for value in header_values if value is not None})
            )
        table_ref = _select_manifest_table_ref(
            table_records,
            sheet_name=actual_name,
            required_fields=(sheet_field, csv_name_field, export_flag_field),
        )
        if table_ref is None:
            rows = [
                (row_number, list(values))
                for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1)
            ]
        else:
            min_column, min_row, max_column, max_row = range_boundaries(table_ref)
            rows = [
                (row_number, list(values))
                for row_number, values in enumerate(
                    sheet.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_column,
                        max_col=max_column,
                        values_only=True,
                    ),
                    start=min_row,
                )
            ]
        return _parse_manifest_rows(
            rows,
            sheet_name=actual_name,
            sheet_field=sheet_field,
            csv_name_field=csv_name_field,
            export_flag_field=export_flag_field,
            parser="openpyxl",
        )
    finally:
        workbook.close()


def _column_index(cell_reference: str) -> int:
    match = _CELL_REF_RE.match(cell_reference)
    if match is None:
        return 0
    result = 0
    for char in match.group(1):
        result = result * 26 + ord(char) - ord("A") + 1
    return result - 1


def _read_shared_strings(archive: ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in archive.namelist():
        return []
    root = ET.fromstring(archive.read(path))
    return [
        "".join(node.text or "" for node in item.iter(f"{{{_SHEET_NS}}}t"))
        for item in root.findall(f"{{{_SHEET_NS}}}si")
    ]


def _resolve_sheet_path(archive: ZipFile, configured_name: str) -> tuple[str, str]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    sheets = workbook_root.find(f"{{{_SHEET_NS}}}sheets")
    records: list[tuple[str, str]] = []
    if sheets is not None:
        for sheet in sheets:
            records.append((sheet.attrib.get("name", ""), sheet.attrib.get(f"{{{_REL_NS}}}id", "")))
    actual_name = _find_sheet_name((name for name, _ in records), configured_name)
    relation_id = next(rel_id for name, rel_id in records if name == actual_name)

    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    target = ""
    for relation in rel_root.findall(f"{{{_PACKAGE_REL_NS}}}Relationship"):
        if relation.attrib.get("Id") == relation_id:
            target = relation.attrib.get("Target", "")
            break
    if not target:
        raise KeyError(f"missing workbook relation: {relation_id}")
    if target.startswith("/"):
        path = target.lstrip("/")
    else:
        path = posixpath.normpath(posixpath.join("xl", target))
    return actual_name, path


def _read_ooxml_table_records(
    archive: ZipFile,
    sheet_root: ET.Element,
    sheet_path: str,
) -> list[tuple[str, set[str]]]:
    table_parts = sheet_root.find(f"{{{_SHEET_NS}}}tableParts")
    if table_parts is None:
        return []
    relation_ids = [
        part.attrib.get(f"{{{_REL_NS}}}id", "")
        for part in table_parts.findall(f"{{{_SHEET_NS}}}tablePart")
    ]
    relation_path = posixpath.join(
        posixpath.dirname(sheet_path),
        "_rels",
        f"{posixpath.basename(sheet_path)}.rels",
    )
    relation_root = ET.fromstring(archive.read(relation_path))
    targets = {
        relation.attrib.get("Id", ""): relation.attrib.get("Target", "")
        for relation in relation_root.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
    }

    records: list[tuple[str, set[str]]] = []
    for relation_id in relation_ids:
        target = targets.get(relation_id, "")
        if not target:
            raise KeyError(f"missing worksheet relation: {relation_id}")
        if target.startswith("/"):
            table_path = target.lstrip("/")
        else:
            table_path = posixpath.normpath(
                posixpath.join(posixpath.dirname(sheet_path), target)
            )
        table_root = ET.fromstring(archive.read(table_path))
        table_columns = table_root.find(f"{{{_SHEET_NS}}}tableColumns")
        fields = (
            {
                column.attrib.get("name", "")
                for column in table_columns.findall(f"{{{_SHEET_NS}}}tableColumn")
            }
            if table_columns is not None
            else set()
        )
        records.append((table_root.attrib.get("ref", ""), fields))
    return records


def _resolve_ooxml_table_bounds(
    table_ref: str | None,
    sheet_data: ET.Element | None,
    *,
    sheet_name: str,
    table_width: int | None,
) -> tuple[int, int, int, int] | None:
    if table_ref is None:
        return None
    if not table_ref.strip():
        raise M2ProcessingError(
            "M2_MANIFEST_FIELD_MISSING",
            "manifest_parse",
            "main Sheet 的 manifest Excel Table 范围无效",
            sheet_name=sheet_name,
            details={"table_ref": table_ref},
        )
    raw_bounds = range_boundaries(table_ref)

    min_column, min_row, max_column, max_row = raw_bounds
    resolved_bounds: tuple[int, int, int, int] | None = None
    if all(bound is not None for bound in raw_bounds):
        resolved_bounds = (
            int(min_column),
            int(min_row),
            int(max_column),
            int(max_row),
        )

    if resolved_bounds is None:
        if min_row is None or max_row is None:
            raise M2ProcessingError(
                "M2_MANIFEST_FIELD_MISSING",
                "manifest_parse",
                "main Sheet 的 manifest Excel Table 缺少可靠行边界",
                sheet_name=sheet_name,
                details={"table_ref": table_ref},
            )
        coordinates: list[tuple[int, int]] = []
        if sheet_data is not None:
            for fallback_row, row_node in enumerate(
                sheet_data.findall(f"{{{_SHEET_NS}}}row"),
                start=1,
            ):
                row_number = int(row_node.attrib.get("r", fallback_row))
                if row_number < min_row or row_number > max_row:
                    continue
                for cell in row_node.findall(f"{{{_SHEET_NS}}}c"):
                    column_number = _column_index(cell.attrib.get("r", "A1")) + 1
                    if min_column is not None and column_number < min_column:
                        continue
                    if max_column is not None and column_number > max_column:
                        continue
                    coordinates.append((column_number, row_number))

        if not coordinates:
            raise M2ProcessingError(
                "M2_MANIFEST_FIELD_MISSING",
                "manifest_parse",
                "main Sheet 的 manifest Excel Table 缺失列边界且无法可靠恢复",
                sheet_name=sheet_name,
                details={"table_ref": table_ref},
            )

        used_columns = [column for column, _ in coordinates]
        resolved_bounds = (
            int(min_column) if min_column is not None else min(used_columns),
            int(min_row),
            int(max_column) if max_column is not None else max(used_columns),
            int(max_row),
        )
        inferred_width = resolved_bounds[2] - resolved_bounds[0] + 1
        if table_width is None or inferred_width != table_width:
            raise M2ProcessingError(
                "M2_MANIFEST_FIELD_MISSING",
                "manifest_parse",
                "main Sheet 的 manifest Excel Table 缺失列边界且无法可靠恢复",
                sheet_name=sheet_name,
                details={"table_ref": table_ref},
            )

    return resolved_bounds


def _parse_with_ooxml(
    raw: bytes,
    *,
    sheet_name: str,
    sheet_field: str,
    csv_name_field: str,
    export_flag_field: str,
) -> WorkbookManifest:
    with ZipFile(BytesIO(raw)) as archive:
        actual_name, sheet_path = _resolve_sheet_path(archive, sheet_name)
        shared_strings = _read_shared_strings(archive)
        root = ET.fromstring(archive.read(sheet_path))
        table_records = _read_ooxml_table_records(archive, root, sheet_path)
        required_fields = (sheet_field, csv_name_field, export_flag_field)
        table_ref = _select_manifest_table_ref(
            table_records,
            sheet_name=actual_name,
            required_fields=required_fields,
        )
        table_width = next(
            (
                len(fields)
                for ref, fields in table_records
                if ref == table_ref and set(required_fields) <= fields
            ),
            None,
        )
        sheet_data = root.find(f"{{{_SHEET_NS}}}sheetData")
        bounds = _resolve_ooxml_table_bounds(
            table_ref,
            sheet_data,
            sheet_name=actual_name,
            table_width=table_width,
        )
        rows: list[tuple[int, list[Any]]] = []
        if sheet_data is not None:
            for row_node in sheet_data.findall(f"{{{_SHEET_NS}}}row"):
                row_number = int(row_node.attrib.get("r", len(rows) + 1))
                if bounds is not None and not bounds[1] <= row_number <= bounds[3]:
                    continue
                cells: dict[int, str] = {}
                for cell in row_node.findall(f"{{{_SHEET_NS}}}c"):
                    absolute_index = _column_index(cell.attrib.get("r", "A1"))
                    if bounds is not None:
                        if not bounds[0] - 1 <= absolute_index <= bounds[2] - 1:
                            continue
                        index = absolute_index - (bounds[0] - 1)
                    else:
                        index = absolute_index
                    cell_type = cell.attrib.get("t", "")
                    if cell_type == "inlineStr":
                        value = "".join(
                            node.text or ""
                            for node in cell.iter(f"{{{_SHEET_NS}}}t")
                        )
                    else:
                        value_node = cell.find(f"{{{_SHEET_NS}}}v")
                        value = value_node.text if value_node is not None and value_node.text is not None else ""
                        if cell_type == "s" and value != "":
                            value = shared_strings[int(value)]
                    cells[index] = value
                width = (
                    bounds[2] - bounds[0] + 1
                    if bounds is not None
                    else max(cells, default=-1) + 1
                )
                rows.append((row_number, [cells.get(index, "") for index in range(width)]))
        return _parse_manifest_rows(
            rows,
            sheet_name=actual_name,
            sheet_field=sheet_field,
            csv_name_field=csv_name_field,
            export_flag_field=export_flag_field,
            parser="ooxml",
        )


def parse_workbook_manifest(
    raw: bytes,
    *,
    sheet_name: str = "main",
    sheet_field: str = "sheetName",
    csv_name_field: str = "tbxName",
    export_flag_field: str = "isExport",
) -> WorkbookManifest:
    """优先使用 openpyxl；样式等问题导致加载失败时只读 OOXML 兜底。"""
    try:
        return _parse_with_openpyxl(
            raw,
            sheet_name=sheet_name,
            sheet_field=sheet_field,
            csv_name_field=csv_name_field,
            export_flag_field=export_flag_field,
        )
    except M2ProcessingError:
        raise
    except Exception:
        try:
            return _parse_with_ooxml(
                raw,
                sheet_name=sheet_name,
                sheet_field=sheet_field,
                csv_name_field=csv_name_field,
                export_flag_field=export_flag_field,
            )
        except M2ProcessingError:
            raise
        except (BadZipFile, KeyError, ET.ParseError, IndexError, ValueError) as exc:
            raise M2ProcessingError(
                "M2_WORKBOOK_PARSE_FAILED",
                "workbook_parse",
                "工作簿无法通过 openpyxl 或 OOXML 解析",
                sheet_name=sheet_name,
                details={"parsers": ["openpyxl", "ooxml"]},
            ) from exc
