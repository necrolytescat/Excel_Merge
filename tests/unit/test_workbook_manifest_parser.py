from io import BytesIO
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

import core.workbook_manifest_parser as manifest_parser
from core.m2_errors import M2ProcessingError


_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _replace_zip_member(raw: bytes, path: str, content: bytes) -> bytes:
    source = BytesIO(raw)
    target = BytesIO()
    with ZipFile(source) as input_archive, ZipFile(target, "w", ZIP_DEFLATED) as output_archive:
        for item in input_archive.infolist():
            output_archive.writestr(
                item,
                content if item.filename == path else input_archive.read(item.filename),
            )
    return target.getvalue()


def _set_formula_cache(raw: bytes, reference: str, value: str) -> bytes:
    with ZipFile(BytesIO(raw)) as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    for cell in root.iter(f"{{{_SHEET_NS}}}c"):
        if cell.attrib.get("r") == reference:
            cell.attrib["t"] = "str"
            value_node = cell.find(f"{{{_SHEET_NS}}}v")
            if value_node is None:
                value_node = ET.SubElement(cell, f"{{{_SHEET_NS}}}v")
            value_node.text = value
            break
    else:
        raise AssertionError(f"missing fixture cell: {reference}")
    return _replace_zip_member(
        raw,
        "xl/worksheets/sheet1.xml",
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _with_broken_styles_and_table_ref(raw: bytes, table_ref: str) -> bytes:
    with ZipFile(BytesIO(raw)) as archive:
        styles_root = ET.fromstring(archive.read("xl/styles.xml"))
        table_root = ET.fromstring(archive.read("xl/tables/table1.xml"))
        sheet_root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))

    fills = styles_root.find(f"{{{_SHEET_NS}}}fills")
    assert fills is not None
    fill = ET.SubElement(fills, f"{{{_SHEET_NS}}}fill")
    gradient = ET.SubElement(fill, f"{{{_SHEET_NS}}}gradientFill")
    for color in ("FF000000", "FFFFFFFF"):
        stop = ET.SubElement(
            gradient,
            f"{{{_SHEET_NS}}}stop",
            {"position": "0"},
        )
        ET.SubElement(stop, f"{{{_SHEET_NS}}}color", {"rgb": color})
    fills.attrib["count"] = str(len(fills))

    table_root.attrib["ref"] = table_ref
    auto_filter = table_root.find(f"{{{_SHEET_NS}}}autoFilter")
    if auto_filter is not None:
        auto_filter.attrib["ref"] = table_ref
    dimension = sheet_root.find(f"{{{_SHEET_NS}}}dimension")
    if dimension is not None:
        sheet_root.remove(dimension)

    raw = _replace_zip_member(
        raw,
        "xl/styles.xml",
        ET.tostring(styles_root, encoding="utf-8", xml_declaration=True),
    )
    raw = _replace_zip_member(
        raw,
        "xl/tables/table1.xml",
        ET.tostring(table_root, encoding="utf-8", xml_declaration=True),
    )
    return _replace_zip_member(
        raw,
        "xl/worksheets/sheet1.xml",
        ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True),
    )


def _table_manifest_workbook(
    rows: list[list[object]],
    *,
    table_ref: str,
    extra_rows: list[list[object]] | None = None,
    second_table: bool = False,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "main"
    for row in rows:
        sheet.append(row)
    sheet.add_table(Table(displayName="Manifest", ref=table_ref))
    if extra_rows:
        for row in extra_rows:
            sheet.append(row)
    if second_table:
        for row_number, values in enumerate(
            [
                ["sheetName", "tbxName", "isExport"],
                ["Other", "Other_Base", 1],
            ],
            start=1,
        ):
            for column_number, value in enumerate(values, start=5):
                sheet.cell(row=row_number, column=column_number, value=value)
        sheet.add_table(Table(displayName="ManifestOther", ref="E1:G2"))
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _minimal_manifest_workbook() -> bytes:
    workbook_xml = """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="main" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""
    relationships = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
   Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
   Target="worksheets/sheet1.xml"/>
</Relationships>"""
    sheet_xml = """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" t="inlineStr"><is><t>sheetName</t></is></c>
      <c r="B1" t="inlineStr"><is><t>tbxName</t></is></c>
      <c r="C1" t="inlineStr"><is><t>isExport</t></is></c>
    </row>
    <row r="2">
      <c r="A2" t="inlineStr"><is><t>Base</t></is></c>
      <c r="B2" t="inlineStr"><is><t>AtlasConfig_Base</t></is></c>
      <c r="C2"><v>1</v></c>
    </row>
  </sheetData>
</worksheet>"""
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", relationships)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buffer.getvalue()


def test_manifest_parser_falls_back_to_ooxml(monkeypatch):
    monkeypatch.setattr(
        manifest_parser,
        "load_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("bad styles")),
    )

    manifest = manifest_parser.parse_workbook_manifest(_minimal_manifest_workbook())

    assert manifest.parser == "ooxml"
    assert [(item.sheet_name, item.tbx_name, item.is_export) for item in manifest.entries] == [
        ("Base", "AtlasConfig_Base", "1")
    ]


@pytest.mark.parametrize("table_ref", ["A1:C2", "1:2"])
def test_manifest_ooxml_fallback_infers_missing_table_boundaries(table_ref):
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
        ],
        table_ref="A1:C2",
        extra_rows=[["Ghost", "GhostCsv", 1]],
    )
    raw = _with_broken_styles_and_table_ref(raw, table_ref)

    with pytest.raises(ValueError, match="could not read stylesheet") as captured:
        manifest_parser.load_workbook(BytesIO(raw))
    assert captured.value.__cause__ is not None
    assert "Duplicate position 0" in str(captured.value.__cause__)

    manifest = manifest_parser.parse_workbook_manifest(raw)

    assert manifest.parser == "ooxml"
    assert [
        (item.sheet_name, item.tbx_name, item.is_export)
        for item in manifest.entries
    ] == [("Base", "AtlasConfig_Base", "1")]


def test_manifest_ooxml_row_bounds_reject_horizontal_cells_outside_table():
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport", "helper"],
            ["Base", "AtlasConfig_Base", 1, "outside table"],
        ],
        table_ref="A1:C2",
        extra_rows=[["Ghost", "GhostCsv", 1]],
    )
    raw = _with_broken_styles_and_table_ref(raw, "1:2")

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {"table_ref": "1:2"}


@pytest.mark.parametrize("table_ref", ["0:2", "1:1048577"])
def test_manifest_ooxml_row_bounds_reject_rows_outside_excel(table_ref):
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
        ],
        table_ref="A1:C2",
        extra_rows=[["Ghost", "GhostCsv", 1]],
    )
    raw = _with_broken_styles_and_table_ref(raw, table_ref)

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {"table_ref": table_ref}


def test_manifest_ooxml_row_bounds_reject_inferred_columns_outside_excel():
    sheet_data = ET.fromstring(
        f"""<sheetData xmlns="{_SHEET_NS}">
  <row r="1"><c r="XFE1"/><c r="XFF1"/><c r="XFG1"/></row>
  <row r="2"><c r="XFE2"/><c r="XFF2"/><c r="XFG2"/></row>
</sheetData>"""
    )

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser._resolve_ooxml_table_bounds(
            "1:2",
            sheet_data,
            sheet_name="main",
            table_width=3,
        )

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {"table_ref": "1:2"}


@pytest.mark.parametrize(
    "table_ref",
    ["", "A:C"],
)
def test_manifest_ooxml_fallback_rejects_unprovable_table_boundary(table_ref):
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
        ],
        table_ref="A1:C2",
        extra_rows=[["Ghost", "GhostCsv", 1]],
    )
    raw = _with_broken_styles_and_table_ref(raw, table_ref)

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {"table_ref": table_ref}


def test_manifest_ooxml_fallback_preserves_malformed_boundary_error_contract():
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
        ],
        table_ref="A1:C2",
    )
    raw = _with_broken_styles_and_table_ref(raw, "invalid-range")

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_WORKBOOK_PARSE_FAILED"
    assert captured.value.details == {"parsers": ["openpyxl", "ooxml"]}


def test_manifest_ooxml_fallback_preserves_reversed_boundary_error_contract():
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
        ],
        table_ref="A1:C2",
    )
    raw = _with_broken_styles_and_table_ref(raw, "C1:A2")

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {"header_match_count": 0}


@pytest.mark.parametrize("force_ooxml", [False, True])
def test_manifest_uses_formula_cache_and_table_bounds(monkeypatch, force_ooxml):
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", '=IF(C2=1,"AtlasConfig_Base","")', 1],
        ],
        table_ref="A1:C2",
        extra_rows=[["helper", "outside table", 1]],
    )
    raw = _set_formula_cache(raw, "B2", "AtlasConfig_Base")
    if force_ooxml:
        monkeypatch.setattr(
            manifest_parser,
            "load_workbook",
            lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("bad styles")),
        )

    manifest = manifest_parser.parse_workbook_manifest(raw)

    assert manifest.parser == ("ooxml" if force_ooxml else "openpyxl")
    assert [(item.sheet_name, item.tbx_name, item.is_export) for item in manifest.entries] == [
        ("Base", "AtlasConfig_Base", "1")
    ]


def test_manifest_only_includes_isexport_one():
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
            ["Statement", "QuestConfig_Statement", 0],
            ["Condition", None, 0],
            ["Notes", None, None],
            ["Unknown", "Unknown_Base", "yes"],
        ],
        table_ref="A1:C6",
    )

    manifest = manifest_parser.parse_workbook_manifest(raw)

    assert [item.sheet_name for item in manifest.entries] == ["Base"]


def test_manifest_rejects_exported_row_without_formula_cache():
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", '=IF(C2=1,"AtlasConfig_Base","")', 1],
        ],
        table_ref="A1:C2",
    )

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {"row": 2, "fields": ["tbxName"]}


def test_manifest_rejects_multiple_candidate_tables():
    raw = _table_manifest_workbook(
        [
            ["sheetName", "tbxName", "isExport"],
            ["Base", "AtlasConfig_Base", 1],
        ],
        table_ref="A1:C2",
        second_table=True,
    )

    with pytest.raises(M2ProcessingError) as captured:
        manifest_parser.parse_workbook_manifest(raw)

    assert captured.value.code == "M2_MANIFEST_FIELD_MISSING"
    assert captured.value.details == {
        "candidate_table_count": 2,
        "table_refs": ["A1:C2", "E1:G2"],
    }
