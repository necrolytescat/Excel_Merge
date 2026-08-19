from __future__ import annotations

from io import BytesIO
import json

import pytest
from openpyxl import load_workbook

from app.schemas.diff_export import DiffExportRequestPayload
from app.services.diff_export_service import DiffExportError, DiffExportService


def _diff_payload(*, source_only: bool = False) -> bytes:
    rows = [
        {
            "key": "1",
            "status": "modified",
            "source": {"row_number": 8, "values": {"Id": "1", "Name": "左侧"}},
            "target": {"row_number": 8, "values": {"Id": "1", "Name": "右侧", "Region": "CN"}},
            "changes": [],
        }
    ]
    if source_only:
        rows.append(
            {
                "key": "2",
                "status": "source_only",
                "source": {"row_number": 9, "values": {"Id": "2", "Name": "新增"}},
                "changes": [],
            }
        )
    else:
        rows.append(
            {
                "key": "2",
                "status": "target_only",
                "target": {"row_number": 9, "values": {"Id": "2", "Name": "删除我", "Region": "CN"}},
                "changes": [],
            }
        )
    payload = {
        "schema_version": "m2.diff.v1",
        "direction": {"source": "left", "target": "right"},
        "workbook": {
            "name": "Demo.xlsx",
            "status": "modified",
            "source_sha256": "source",
            "target_sha256": "target",
        },
        "summary": {"total_sheets": 1, "modified_sheets": 1, "modified_rows": 2},
        "sheets": [
            {
                "sheet_name": "Main",
                "status": "modified",
                "primary_key": "Id",
                "fields": [
                    {
                        "name": "Id",
                        "status": "common",
                        "source_display_name": "编号",
                        "target_display_name": "编号",
                        "source_type": "int",
                        "target_type": "int",
                        "source_scope": "client",
                        "target_scope": "client",
                    },
                    {
                        "name": "Name",
                        "status": "common",
                        "source_display_name": "名称",
                        "target_display_name": "名称",
                        "source_type": "string",
                        "target_type": "string",
                        "source_scope": "client",
                        "target_scope": "client",
                    },
                    {
                        "name": "Region",
                        "status": "target_only",
                        "target_display_name": "区域",
                        "target_type": "string",
                        "target_scope": "client",
                    },
                ],
                "rows": rows,
                "errors": [],
            }
        ],
        "errors": [],
    }
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def _request(*decisions: dict) -> DiffExportRequestPayload:
    return DiffExportRequestPayload.model_validate(
        {
            "target_layout": "target",
            "sheets": [{"sheet_name": "Main", "decisions": list(decisions)}],
        }
    )


def test_export_mixes_sources_and_marks_fallback_cells() -> None:
    artifact = DiffExportService(lambda _: (_diff_payload(), "sha")).export(
        "m2r_test",
        _request(
            {"key": "1", "action": "write", "value_side": "source"},
            {"key": "2", "action": "delete"},
        ),
    )

    workbook = load_workbook(BytesIO(artifact.content))
    assert workbook.sheetnames == ["导出说明", "Main"]
    sheet = workbook["Main"]
    assert [sheet.cell(1, column).value for column in range(1, 5)] == ["操作", "编号", "名称", "区域"]
    assert [sheet.cell(2, column).value for column in range(1, 5)] == ["__action__", "Id", "Name", "Region"]
    assert [sheet.cell(3, column).value for column in range(1, 5)] == ["写入", "1", "左侧", "CN"]
    assert sheet["D3"].fill.fgColor.rgb.endswith("FF0000")
    assert sheet["D3"].font.color.rgb.endswith("FFFFFF")
    assert [sheet.cell(4, column).value for column in range(1, 5)] == ["删除", "2", "删除我", "CN"]
    assert sheet.freeze_panes == "B3"
    assert sheet.auto_filter.ref == "A2:D4"


def test_export_blocks_source_only_row_when_target_field_cannot_be_filled() -> None:
    with pytest.raises(DiffExportError) as raised:
        DiffExportService(lambda _: (_diff_payload(source_only=True), "sha")).export(
            "m2r_test",
            _request({"key": "2", "action": "write", "value_side": "source"}),
        )

    assert raised.value.code == "DIFF_EXPORT_VALIDATION_FAILED"
    assert {issue.code for issue in raised.value.issues} == {"EXPORT_TARGET_FIELD_MISSING"}


def test_export_rejects_delete_for_modified_row() -> None:
    with pytest.raises(DiffExportError) as raised:
        DiffExportService(lambda _: (_diff_payload(), "sha")).export(
            "m2r_test",
            _request({"key": "1", "action": "delete"}),
        )

    assert {issue.code for issue in raised.value.issues} == {"EXPORT_DELETE_NOT_ALLOWED"}


def test_export_writes_multiple_selected_sheets_and_omits_empty_sheet() -> None:
    payload = json.loads(_diff_payload().decode("utf-8"))
    second = json.loads(json.dumps(payload["sheets"][0]))
    second["sheet_name"] = "Second"
    second["rows"][0]["key"] = "3"
    second["rows"][0]["source"]["values"]["Id"] = "3"
    second["rows"][0]["target"]["values"]["Id"] = "3"
    payload["sheets"].append(second)
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    request = DiffExportRequestPayload.model_validate(
        {
            "target_layout": "target",
            "sheets": [
                {"sheet_name": "Main", "decisions": [{"key": "1", "action": "write", "value_side": "target"}]},
                {"sheet_name": "Second", "decisions": [{"key": "3", "action": "write", "value_side": "source"}]},
            ],
        }
    )
    artifact = DiffExportService(lambda _: (raw, "sha")).export("m2r_test", request)
    workbook = load_workbook(BytesIO(artifact.content))
    assert workbook.sheetnames == ["导出说明", "Main", "Second"]
    assert workbook["Main"]["C3"].value == "右侧"
    assert workbook["Second"]["C3"].value == "左侧"
