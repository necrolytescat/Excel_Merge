from __future__ import annotations

from io import BytesIO
import json

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import create_app
from core.svn_provider import MockSVNProvider


class FakeBatchService:
    def __init__(self, content: bytes):
        self.content = content

    def load_result(self, result_ref: str):
        assert result_ref == "m2r_fixture"
        return self.content, "fixture-sha"

    def close(self):
        return None


def diff_result() -> bytes:
    return json.dumps(
        {
            "schema_version": "m2.diff.v1",
            "direction": {"source": "left", "target": "right"},
            "workbook": {
                "name": "Config.xlsx",
                "status": "modified",
                "source_sha256": "s",
                "target_sha256": "t",
            },
            "summary": {"total_sheets": 1, "modified_sheets": 1, "modified_rows": 1},
            "sheets": [
                {
                    "sheet_name": "Main",
                    "status": "modified",
                    "primary_key": "Id",
                    "fields": [
                        {
                            "name": "Id",
                            "status": "common",
                            "source_display_name": "编号",
                            "target_display_name": "编号",
                            "source_type": "int",
                            "target_type": "int",
                        },
                        {
                            "name": "Name",
                            "status": "common",
                            "source_display_name": "名称",
                            "target_display_name": "名称",
                            "source_type": "string",
                            "target_type": "string",
                        },
                    ],
                    "rows": [
                        {
                            "key": "1",
                            "status": "modified",
                            "source": {"row_number": 8, "values": {"Id": "1", "Name": "左侧"}},
                            "target": {"row_number": 8, "values": {"Id": "1", "Name": "右侧"}},
                            "changes": [],
                        }
                    ],
                    "errors": [],
                }
            ],
            "errors": [],
        },
        ensure_ascii=False,
    ).encode("utf-8")


def test_export_endpoint_returns_xlsx_without_touching_result() -> None:
    app = create_app(
        config={"svn": {"provider": "mock"}},
        provider=MockSVNProvider(),
        batch_diff_service=FakeBatchService(diff_result()),
    )
    with TestClient(app) as client:
        response = client.post(
            "/api/diff/batch-results/m2r_fixture/export",
            json={
                "schema_version": "m2.export.v1",
                "target_layout": "target",
                "sheets": [
                    {
                        "sheet_name": "Main",
                        "decisions": [{"key": "1", "action": "write", "value_side": "source"}],
                    }
                ],
            },
        )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["导出说明", "Main"]
    assert workbook["Main"]["C3"].value == "左侧"


def test_export_endpoint_returns_structured_validation_error() -> None:
    app = create_app(
        config={"svn": {"provider": "mock"}},
        provider=MockSVNProvider(),
        batch_diff_service=FakeBatchService(diff_result()),
    )
    with TestClient(app) as client:
        response = client.post(
            "/api/diff/batch-results/m2r_fixture/export",
            json={
                "schema_version": "m2.export.v1",
                "target_layout": "target",
                "sheets": [
                    {
                        "sheet_name": "Main",
                        "decisions": [{"key": "missing", "action": "write", "value_side": "source"}],
                    }
                ],
            },
        )

    assert response.status_code == 422
    body = response.json()["error"]
    assert body["code"] == "DIFF_EXPORT_VALIDATION_FAILED"
    assert body["details"]["schema_version"] == "m2.export-validation.v1"
    assert body["details"]["issues"][0]["code"] == "EXPORT_ROW_NOT_FOUND"
