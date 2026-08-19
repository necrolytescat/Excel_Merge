"""Server-side XLSX export derived from persisted m2.diff.v1 results."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import json
from pathlib import PurePosixPath
import re
from typing import Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.diff import DiffResultPayload, FieldDefinitionPayload, RowDiffPayload, SheetDiffPayload
from app.schemas.diff_export import (
    DiffExportRequestPayload,
    DiffExportSummaryPayload,
    ExportAction,
    ExportRowDecision,
    ExportSheetSummary,
    ExportSide,
    ExportValidationIssue,
)


class DiffExportError(Exception):
    def __init__(self, code: str, message: str, *, status_code: int, issues: list[ExportValidationIssue] | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code
        self.issues = issues or []


@dataclass(frozen=True)
class DiffExportArtifact:
    content: bytes
    filename: str
    summary: DiffExportSummaryPayload


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_ACTION_DISPLAY_NAME = "操作"
_ACTION_FIELD_NAME = "__action__"
_FALLBACK_FILL = PatternFill(fill_type="solid", fgColor="FF0000")
_FALLBACK_FONT = Font(color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_DELETE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


class DiffExportService:
    def __init__(self, result_loader: Callable[[str], tuple[bytes, str]]):
        self._result_loader = result_loader

    def export(self, result_ref: str, request: DiffExportRequestPayload) -> DiffExportArtifact:
        raw, _sha256 = self._result_loader(result_ref)
        try:
            diff = DiffResultPayload.model_validate_json(raw)
        except Exception as exc:  # pragma: no cover - corrupt result is guarded by BatchStore
            raise DiffExportError(
                "DIFF_EXPORT_RESULT_INVALID",
                "批量结果不是合法的 m2.diff.v1",
                status_code=500,
            ) from exc

        issues: list[ExportValidationIssue] = []
        sheet_map = {sheet.sheet_name: sheet for sheet in diff.sheets}
        if diff.workbook.status.value not in {"modified", "partial"}:
            issues.append(ExportValidationIssue(code="EXPORT_WORKBOOK_NOT_EXPORTABLE", message="当前工作簿没有可导出的语义差异。"))

        if len({item.sheet_name for item in request.sheets}) != len(request.sheets):
            issues.append(ExportValidationIssue(code="EXPORT_DUPLICATE_SHEET", message="导出请求中存在重复 Sheet。"))

        selected: list[tuple[SheetDiffPayload, list[ExportRowDecision]]] = []
        for sheet_request in request.sheets:
            sheet = sheet_map.get(sheet_request.sheet_name)
            if sheet is None:
                issues.append(ExportValidationIssue(
                    code="EXPORT_SHEET_NOT_FOUND",
                    message="请求的 Sheet 不存在于 Diff 结果。",
                    sheet_name=sheet_request.sheet_name,
                ))
                continue
            if sheet.status.value == "failed":
                issues.append(ExportValidationIssue(
                    code="EXPORT_SHEET_FAILED",
                    message="该 Sheet 的 Diff 结果失败，不能导出。",
                    sheet_name=sheet.sheet_name,
                ))
                continue
            selected.append((sheet, sheet_request.decisions))

        if not selected:
            issues.append(ExportValidationIssue(code="EXPORT_NO_SELECTION", message="至少选择一个包含行决策的 Sheet。"))

        prepared: list[tuple[SheetDiffPayload, list[ExportRowDecision], list[str], dict[str, str]]] = []
        for sheet, decisions in selected:
            row_map = {row.key: row for row in sheet.rows}
            seen: set[str] = set()
            fields = self._target_fields(sheet, request.target_layout)
            for decision in decisions:
                if decision.key in seen:
                    issues.append(ExportValidationIssue(
                        code="EXPORT_DUPLICATE_ROW_DECISION",
                        message="同一主键只能有一个导出决策。",
                        sheet_name=sheet.sheet_name,
                        key=decision.key,
                    ))
                seen.add(decision.key)
                row = row_map.get(decision.key)
                if row is None:
                    issues.append(ExportValidationIssue(
                        code="EXPORT_ROW_NOT_FOUND",
                        message="请求的主键不存在于 Diff 结果。",
                        sheet_name=sheet.sheet_name,
                        key=decision.key,
                    ))
                    continue
                self._validate_decision(sheet, row, decision, request.target_layout, fields, issues)
            prepared.append((sheet, decisions, fields, self._header_names(sheet, request.target_layout, fields)))

        if issues:
            raise DiffExportError(
                "DIFF_EXPORT_VALIDATION_FAILED",
                "导出校验失败，请修正全部问题后重试。",
                status_code=422,
                issues=issues,
            )

        workbook = Workbook()
        workbook.remove(workbook.active)
        info = workbook.create_sheet("导出说明")
        used_names = {info.title}
        summaries: list[ExportSheetSummary] = []
        reports: list[dict[str, object]] = []
        for sheet, decisions, fields, headers in prepared:
            output_name = self._safe_sheet_name(sheet.sheet_name, used_names)
            output = workbook.create_sheet(output_name)
            generated = self._write_data_sheet(output, sheet, decisions, request.target_layout, fields, headers)
            sheet_summary = {key: value for key, value in generated.items() if key != "reports"}
            reports.extend(
                {
                    "sheet_name": sheet.sheet_name,
                    "output_sheet_name": output_name,
                    **report,
                }
                for report in generated["reports"]
            )
            summaries.append(ExportSheetSummary(
                sheet_name=sheet.sheet_name,
                output_sheet_name=output_name,
                **sheet_summary,
            ))
        self._write_info_sheet(info, diff, request, summaries, reports)

        summary_payload = DiffExportSummaryPayload(
            target_layout=request.target_layout,
            workbook_name=diff.workbook.name,
            sheets=summaries,
        )
        # Keep workbook metadata deterministic and prevent Excel from treating values as formulas.
        workbook.properties.creator = "Excel Diff/Merge"
        workbook.properties.title = f"差异导出 · {diff.workbook.name}"
        workbook.properties.subject = "m2.diff.v1 selected row export"
        buffer = BytesIO()
        workbook.save(buffer)
        filename = f"{self._workbook_stem(diff.workbook.name)}-差异导出.xlsx"
        return DiffExportArtifact(content=buffer.getvalue(), filename=filename, summary=summary_payload)

    @staticmethod
    def _target_fields(sheet: SheetDiffPayload, target_layout: ExportSide) -> list[str]:
        fields: list[str] = []
        for definition in sheet.fields:
            present = (
                definition.source_display_name is not None
                or definition.source_type is not None
                or definition.source_scope is not None
            ) if target_layout is ExportSide.SOURCE else (
                definition.target_display_name is not None
                or definition.target_type is not None
                or definition.target_scope is not None
            )
            if present and definition.name not in fields:
                fields.append(definition.name)
        if sheet.primary_key and sheet.primary_key not in fields:
            fields.insert(0, sheet.primary_key)
        return fields

    @staticmethod
    def _header_names(sheet: SheetDiffPayload, target_layout: ExportSide, fields: list[str]) -> dict[str, str]:
        definitions = {definition.name: definition for definition in sheet.fields}
        names: dict[str, str] = {}
        for field in fields:
            definition = definitions.get(field)
            if definition is None:
                names[field] = field
                continue
            display = definition.source_display_name if target_layout is ExportSide.SOURCE else definition.target_display_name
            names[field] = display or field
        return names

    def _validate_decision(
        self,
        sheet: SheetDiffPayload,
        row: RowDiffPayload,
        decision: ExportRowDecision,
        target_layout: ExportSide,
        fields: list[str],
        issues: list[ExportValidationIssue],
    ) -> None:
        source_exists = row.source is not None
        target_exists = row.target is not None
        if decision.action is ExportAction.DELETE:
            if not target_exists or source_exists:
                issues.append(ExportValidationIssue(
                    code="EXPORT_DELETE_NOT_ALLOWED",
                    message="只有目标侧独有行才能标记删除。",
                    sheet_name=sheet.sheet_name,
                    key=row.key,
                ))
            return
        if decision.value_side is None:
            issues.append(ExportValidationIssue(
                code="EXPORT_VALUE_SIDE_REQUIRED",
                message="写入决策必须指定左侧或右侧数据来源。",
                sheet_name=sheet.sheet_name,
                key=row.key,
            ))
            return
        values = row.source.values if decision.value_side is ExportSide.SOURCE and row.source else None
        if values is None:
            values = row.target.values if decision.value_side is ExportSide.TARGET and row.target else None
        if values is None:
            issues.append(ExportValidationIssue(
                code="EXPORT_VALUE_SIDE_ROW_MISSING",
                message="所选数据来源不存在该行。",
                sheet_name=sheet.sheet_name,
                key=row.key,
            ))
            return
        target_values = row.source.values if target_layout is ExportSide.SOURCE and row.source else (
            row.target.values if target_layout is ExportSide.TARGET and row.target else None
        )
        definitions = {definition.name: definition for definition in sheet.fields}
        for field in fields:
            if field in values:
                value = values[field]
                definition = definitions.get(field)
                target_type = (
                    definition.source_type if target_layout is ExportSide.SOURCE else definition.target_type
                ) if definition else None
                if not self._value_matches_type(value, target_type):
                    issues.append(ExportValidationIssue(
                        code="EXPORT_VALUE_TYPE_INVALID",
                        message="所选值无法按目标字段类型解释。",
                        sheet_name=sheet.sheet_name,
                        key=row.key,
                        field=field,
                        details={"target_type": target_type, "value": value},
                    ))
                continue
            if target_values is None or field not in target_values:
                issues.append(ExportValidationIssue(
                    code="EXPORT_TARGET_FIELD_MISSING",
                    message="所选侧没有该目标字段，且目标行没有可保留的原值。",
                    sheet_name=sheet.sheet_name,
                    key=row.key,
                    field=field,
                ))

    @staticmethod
    def _value_matches_type(value: str, declared_type: str | None) -> bool:
        if not declared_type or value == "":
            return True
        kind = declared_type.casefold().strip()
        try:
            if kind in {"int", "integer", "long"}:
                int(value)
            elif kind in {"float", "double", "decimal", "number"}:
                float(value)
            elif kind in {"bool", "boolean"}:
                if value.casefold() not in {"true", "false", "0", "1"}:
                    return False
            elif kind in {"date"}:
                date.fromisoformat(value)
            elif kind in {"datetime", "timestamp"}:
                datetime.fromisoformat(value.replace("Z", "+00:00"))
        except (TypeError, ValueError):
            return False
        return True

    def _write_data_sheet(
        self,
        worksheet: Worksheet,
        sheet: SheetDiffPayload,
        decisions: list[ExportRowDecision],
        target_layout: ExportSide,
        fields: list[str],
        headers: dict[str, str],
    ) -> dict[str, int]:
        worksheet.cell(row=1, column=1, value=_ACTION_DISPLAY_NAME)
        worksheet.cell(row=2, column=1, value=_ACTION_FIELD_NAME)
        for index, field in enumerate(fields, start=2):
            worksheet.cell(row=1, column=index, value=headers[field])
            worksheet.cell(row=2, column=index, value=field)
        for row in (1, 2):
            for cell in worksheet[row][: len(fields) + 1]:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

        row_map = {row.key: row for row in sheet.rows}
        write_count = 0
        delete_count = 0
        fallback_count = 0
        omitted_field_count = 0
        reports: list[dict[str, object]] = []
        output_row = 3
        definitions = {definition.name: definition for definition in sheet.fields}
        target_field_set = set(fields)
        for decision in decisions:
            diff_row = row_map[decision.key]
            if decision.action is ExportAction.DELETE:
                values = diff_row.target.values if diff_row.target else {}
                action = "删除"
                delete_count += 1
            else:
                action = "写入"
                write_count += 1
                values = (
                    diff_row.source.values if decision.value_side is ExportSide.SOURCE and diff_row.source else
                    diff_row.target.values if diff_row.target else {}
                )
            worksheet.cell(row=output_row, column=1, value=action)
            if action == "删除":
                worksheet.cell(row=output_row, column=1).fill = _DELETE_FILL
            target_values = (
                diff_row.target.values if decision.action is ExportAction.DELETE and diff_row.target else
                diff_row.source.values if target_layout is ExportSide.SOURCE and diff_row.source else
                diff_row.target.values if diff_row.target else {}
            )
            fallback_fields: list[str] = []
            for index, field in enumerate(fields, start=2):
                cell = worksheet.cell(row=output_row, column=index)
                if field in values:
                    cell.value = self._safe_cell_value(values[field])
                elif field in target_values:
                    cell.value = self._safe_cell_value(target_values[field])
                    cell.fill = _FALLBACK_FILL
                    cell.font = _FALLBACK_FONT
                    fallback_fields.append(field)
                    fallback_count += 1
                else:
                    cell.value = ""
                cell.alignment = Alignment(vertical="top", wrap_text=False)
            omitted_fields = sorted(set(values) - target_field_set)
            omitted_field_count += len(omitted_fields)
            type_warning_fields: list[str] = []
            for field in fields:
                definition = definitions.get(field)
                if not definition or definition.status.value != "modified" or field not in values:
                    continue
                chosen_type = definition.source_type if decision.value_side is ExportSide.SOURCE else definition.target_type
                target_type = definition.source_type if target_layout is ExportSide.SOURCE else definition.target_type
                if chosen_type and target_type and chosen_type != target_type:
                    type_warning_fields.append(field)
            reports.append({
                "key": diff_row.key,
                "action": action,
                "value_side": "目标侧" if action == "删除" else ("左侧" if decision.value_side is ExportSide.SOURCE else "右侧"),
                "target_row_number": diff_row.target.row_number if diff_row.target else "",
                "fallback_fields": fallback_fields,
                "omitted_fields": omitted_fields,
                "type_warning_fields": type_warning_fields,
            })
            output_row += 1

        worksheet.freeze_panes = "B3"
        worksheet.auto_filter.ref = f"A2:{get_column_letter(len(fields) + 1)}{max(2, output_row - 1)}"
        worksheet.sheet_view.showGridLines = False
        widths = {1: 10}
        for index, field in enumerate(fields, start=2):
            widths[index] = min(40, max(12, len(headers[field]) + 2, len(field) + 2))
        for column, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column)].width = width
        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[2].height = 22
        return {
            "write_count": write_count,
            "delete_count": delete_count,
            "fallback_count": fallback_count,
            "omitted_field_count": omitted_field_count,
            "reports": reports,
        }

    @staticmethod
    def _write_info_sheet(
        worksheet: Worksheet,
        diff: DiffResultPayload,
        request: DiffExportRequestPayload,
        summaries: list[ExportSheetSummary],
        reports: list[dict[str, object]],
    ) -> None:
        rows: list[list[str | int]] = [
            ["差异导出说明", ""],
            ["工作簿", diff.workbook.name],
            ["目标结构", "左侧" if request.target_layout is ExportSide.SOURCE else "右侧"],
            ["红底白字", "所选侧缺少字段，保留目标侧原值"],
            ["复制写入", "筛选 操作=写入，从 B 列开始复制"],
            ["删除处理", "筛选 操作=删除，在目标 Sheet 中人工删除对应主键行"],
            [],
            ["原 Sheet", "输出 Sheet", "写入行数", "删除行数", "目标侧补位字段数", "省略字段计数"],
        ]
        for summary in summaries:
            rows.append([
                summary.sheet_name,
                summary.output_sheet_name,
                summary.write_count,
                summary.delete_count,
                summary.fallback_count,
                summary.omitted_field_count,
            ])
        rows.extend([
            [],
            ["逐行来源与风险"],
            ["原 Sheet", "输出 Sheet", "主键", "操作", "数据来源", "目标行号", "补位字段", "省略字段", "类型警告"],
        ])
        for report in reports:
            rows.append([
                report["sheet_name"],
                report["output_sheet_name"],
                report["key"],
                report["action"],
                report["value_side"],
                report["target_row_number"],
                ", ".join(report["fallback_fields"]),
                ", ".join(report["omitted_fields"]),
                ", ".join(report["type_warning_fields"]),
            ])
        for row in rows:
            worksheet.append(row)
        for row_index in (1, 8, 10):
            for cell in worksheet[row_index]:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
        worksheet.freeze_panes = "A11"
        worksheet.sheet_view.showGridLines = False
        for index, width in enumerate((24, 28, 18, 12, 12, 12, 22, 22, 22), start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

    @staticmethod
    def _safe_cell_value(value: str) -> str:
        # Export values as strings to preserve the original displayed value and avoid formula execution.
        text = "" if value is None else str(value)
        return text

    @staticmethod
    def _workbook_stem(name: str) -> str:
        stem = PurePosixPath(str(name).replace("\\", "/")).name
        return re.sub(r"\.(?:xlsx|xlsm|xls)$", "", stem, flags=re.IGNORECASE) or "差异结果"

    @staticmethod
    def _safe_sheet_name(name: str, used: set[str]) -> str:
        base = _INVALID_SHEET_CHARS.sub("_", str(name)).strip() or "Sheet"
        base = base[:31]
        candidate = base
        index = 1
        while candidate.casefold() in {item.casefold() for item in used}:
            suffix = f"~{index}"
            candidate = f"{base[:31 - len(suffix)]}{suffix}"
            index += 1
        used.add(candidate)
        return candidate

